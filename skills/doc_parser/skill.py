#!/usr/bin/env python3
"""Document Parser - pdfplumber + python-docx + openpyxl (MIT) - 100% Free"""
import json, sys, subprocess, os


class Skill:
    def __init__(self, config: dict = None):
        self.config = config or {}

    def run(self, args: dict) -> dict:
        fp = args.get("file", "")
        if not fp or not os.path.exists(fp):
            return {"error": "file required"}
        ext = os.path.splitext(fp)[1].lower()
        try:
            if ext == ".pdf":
                import pdfplumber
                pages = []
                with pdfplumber.open(fp) as pdf:
                    for p in pdf.pages:
                        pages.append({"page": len(pages)+1, "text": p.extract_text() or ""})
                return {"type": "pdf", "pages": len(pages), "content": pages}
            elif ext == ".docx":
                from docx import Document
                doc = Document(fp)
                return {"type": "docx", "paragraphs": [p.text for p in doc.paragraphs if p.text.strip()]}
            elif ext in (".xlsx", ".xlsm"):
                import openpyxl
                wb = openpyxl.load_workbook(fp)
                sheets = {}
                for name in wb.sheetnames:
                    ws = wb[name]
                    sheets[name] = [[c.value for c in row] for row in ws.iter_rows(max_row=100)]
                return {"type": "xlsx", "sheets": list(sheets.keys()), "data": sheets}
            return {"error": f"unsupported: {ext}"}
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "-q", "pdfplumber", "python-docx", "openpyxl"], capture_output=True)
            return self.run(args)
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))

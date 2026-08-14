#!/usr/bin/env python3
"""Generate all new EvolvixOS skills from GitHub research - all 100% free."""
import os, json

SKILLS = [
    ("data_analyzer", "Data Analyzer", "Analyze CSV/JSON data with SQL. Free, local.", "pandas duckdb",
'''#!/usr/bin/env python3
"""Data Analyzer - Pandas + DuckDB (BSD/MIT) - 100% Free"""
import json, sys, subprocess


class Skill:
    def __init__(self, config: dict = None):
        self.config = config or {}

    def run(self, args: dict) -> dict:
        action = args.get("action", "stats")
        try:
            import pandas as pd
            import duckdb
            if action == "load_csv":
                df = pd.read_csv(args["file"])
                return {"rows": len(df), "columns": list(df.columns), "head": df.head(10).to_dict("records")}
            elif action == "sql":
                df = pd.read_csv(args.get("file", "")) if args.get("file") else pd.DataFrame(args.get("data", []))
                result = duckdb.query(args["query"]).df()
                return {"rows": len(result), "data": result.to_dict("records")}
            elif action == "stats":
                df = pd.read_csv(args["file"]) if args.get("file") else pd.DataFrame(args.get("data", []))
                return {"describe": df.describe().to_dict(), "rows": len(df), "cols": len(df.columns)}
            return {"error": f"unknown: {action}"}
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "-q", "pandas", "duckdb"], capture_output=True)
            return self.run(args)
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),

    ("doc_parser", "Document Parser", "Parse PDF, Word, Excel. Free, local.", "pdfplumber python-docx openpyxl",
'''#!/usr/bin/env python3
"""Document Parser - pdfplumber + python-docx + openpyxl (MIT) - 100% Free"""
import json, sys, subprocess, os


class Skill:
    def __init__(self, config: dict = None):
        self.config = config or {}

    def run(self, args: dict) -> dict:
        fp = args.get("file", "")
        if not fp or not os.path.exists(fp):
            return {"error": "file required"}
        ext = os.path.splitext(fp)[1].lower()
        try:
            if ext == ".pdf":
                import pdfplumber
                pages = []
                with pdfplumber.open(fp) as pdf:
                    for p in pdf.pages:
                        pages.append({"page": len(pages)+1, "text": p.extract_text() or ""})
                return {"type": "pdf", "pages": len(pages), "content": pages}
            elif ext == ".docx":
                from docx import Document
                doc = Document(fp)
                return {"type": "docx", "paragraphs": [p.text for p in doc.paragraphs if p.text.strip()]}
            elif ext in (".xlsx", ".xlsm"):
                import openpyxl
                wb = openpyxl.load_workbook(fp)
                sheets = {}
                for name in wb.sheetnames:
                    ws = wb[name]
                    sheets[name] = [[c.value for c in row] for row in ws.iter_rows(max_row=100)]
                return {"type": "xlsx", "sheets": list(sheets.keys()), "data": sheets}
            return {"error": f"unsupported: {ext}"}
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "-q", "pdfplumber", "python-docx", "openpyxl"], capture_output=True)
            return self.run(args)
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),

    ("offline_translator", "Offline Translator", "Translate 30+ languages offline. Free.", "argostranslate",
'''#!/usr/bin/env python3
"""Offline Translator - Argos Translate (MIT) - 100% Free"""
import json, sys, subprocess


class Skill:
    def __init__(self, config: dict = None):
        self.config = config or {}

    def run(self, args: dict) -> dict:
        text = args.get("text", "")
        if not text:
            return {"error": "text required"}
        try:
            from argostranslate import translate
            installed = translate.get_installed_languages()
            src = [l for l in installed if l.code == args.get("from", "en")]
            dst = [l for l in installed if l.code == args.get("to", "es")]
            if src and dst:
                return {"translated": src[0].get_translation(dst[0]).translate(text)}
            return {"error": "Language pair not installed"}
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "-q", "argostranslate"], capture_output=True)
            return self.run(args)
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),

    ("ocr_scanner", "OCR Scanner", "Extract text from images. Free, local.", "easyocr",
'''#!/usr/bin/env python3
"""OCR Scanner - EasyOCR (Apache 2.0) - 100% Free"""
import json, sys, subprocess, os


class Skill:
    def __init__(self, config: dict = None):
        self.config = config or {}

    def run(self, args: dict) -> dict:
        img = args.get("image", "")
        if not img or not os.path.exists(img):
            return {"error": "image required"}
        try:
            import easyocr
            reader = easyocr.Reader(args.get("languages", ["en"]))
            results = reader.readtext(img)
            return {"text": " ".join([r[1] for r in results]), "regions": len(results)}
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "-q", "easyocr"], capture_output=True)
            return self.run(args)
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),

    ("image_processor", "Image Processor", "Resize, crop, convert, filter images. Free.", "Pillow opencv-python-headless",
'''#!/usr/bin/env python3
"""Image Processor - Pillow + OpenCV (HPND/Apache 2.0) - 100% Free"""
import json, sys, subprocess, os


class Skill:
    def __init__(self, config: dict = None):
        self.config = config or {}

    def run(self, args: dict) -> dict:
        action = args.get("action", "info")
        path = args.get("file", "")
        try:
            from PIL import Image, ImageFilter
            if action == "info":
                img = Image.open(path)
                return {"size": img.size, "mode": img.mode, "format": img.format}
            elif action == "resize":
                img = Image.open(path)
                w, h = args.get("width"), args.get("height")
                if w and h: img = img.resize((w, h))
                elif w: img = img.resize((w, int(img.height * w / img.width)))
                out = args.get("output", path.replace(".", "_resized."))
                img.save(out)
                return {"output": out, "size": img.size}
            elif action == "convert":
                img = Image.open(path)
                fmt = args.get("format", "PNG")
                out = os.path.splitext(path)[0] + "." + fmt.lower()
                img.save(out, format=fmt)
                return {"output": out}
            elif action == "filter":
                img = Image.open(path)
                f = args.get("filter", "blur")
                if f == "grayscale": img = img.convert("L")
                elif f == "blur": img = img.filter(ImageFilter.Blur())
                elif f == "sharpen": img = img.filter(ImageFilter.SHARPEN())
                elif f == "edge": img = img.filter(ImageFilter.FIND_EDGES())
                out = args.get("output", path.replace(".", f"_{f}."))
                img.save(out)
                return {"output": out}
            return {"error": f"unknown: {action}"}
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "-q", "Pillow", "opencv-python-headless"], capture_output=True)
            return self.run(args)
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),

    ("audio_editor", "Audio Editor", "Cut, merge, convert audio. Free, local.", "pydub",
'''#!/usr/bin/env python3
"""Audio Editor - pydub (MIT) - 100% Free"""
import json, sys, subprocess, os


class Skill:
    def __init__(self, config: dict = None):
        self.config = config or {}

    def run(self, args: dict) -> dict:
        action = args.get("action", "info")
        try:
            from pydub import AudioSegment
            if action == "info":
                a = AudioSegment.from_file(args["file"])
                return {"duration_sec": len(a)/1000, "channels": a.channels, "sample_rate": a.frame_rate}
            elif action == "cut":
                a = AudioSegment.from_file(args["file"])
                clip = a[args.get("start_ms",0):args.get("end_ms",len(a))]
                out = args.get("output", "clip.wav")
                clip.export(out, format="wav")
                return {"output": out, "duration_sec": len(clip)/1000}
            elif action == "merge":
                segments = [AudioSegment.from_file(f) for f in args["files"]]
                combined = sum(segments)
                out = args.get("output", "merged.wav")
                combined.export(out, format="wav")
                return {"output": out}
            elif action == "convert":
                a = AudioSegment.from_file(args["file"])
                fmt = args.get("format", "mp3")
                out = os.path.splitext(args["file"])[0] + "." + fmt
                a.export(out, format=fmt)
                return {"output": out}
            return {"error": f"unknown: {action}"}
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "-q", "pydub"], capture_output=True)
            return self.run(args)
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),

    ("video_editor", "Video Editor", "Cut, merge, extract audio from video. Free.", "moviepy",
'''#!/usr/bin/env python3
"""Video Editor - MoviePy (MIT) - 100% Free"""
import json, sys, subprocess


class Skill:
    def __init__(self, config: dict = None):
        self.config = config or {}

    def run(self, args: dict) -> dict:
        action = args.get("action", "info")
        try:
            from moviepy import VideoFileClip, concatenate_videoclips
            if action == "info":
                clip = VideoFileClip(args["file"])
                info = {"duration": clip.duration, "size": clip.size, "fps": clip.fps}
                clip.close()
                return info
            elif action == "cut":
                clip = VideoFileClip(args["file"]).subclip(args["start"], args["end"])
                out = args.get("output", "clip.mp4")
                clip.write_videofile(out, codec="libx264", audio_codec="aac", verbose=False, logger=None)
                clip.close()
                return {"output": out}
            elif action == "merge":
                clips = [VideoFileClip(f) for f in args["files"]]
                final = concatenate_videoclips(clips)
                out = args.get("output", "merged.mp4")
                final.write_videofile(out, codec="libx264", audio_codec="aac", verbose=False, logger=None)
                for c in clips: c.close()
                final.close()
                return {"output": out}
            elif action == "extract_audio":
                clip = VideoFileClip(args["file"])
                out = args.get("output", "audio.wav")
                clip.audio.write_audiofile(out, verbose=False, logger=None)
                clip.close()
                return {"output": out}
            return {"error": f"unknown: {action}"}
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "-q", "moviepy"], capture_output=True)
            return self.run(args)
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),

    ("code_linter", "Code Linter", "Lint and format Python code. Free, local.", "ruff",
'''#!/usr/bin/env python3
"""Code Linter - Ruff (MIT) - 100% Free"""
import json, sys, subprocess, os, tempfile


class Skill:
    def __init__(self, config: dict = None):
        self.config = config or {}

    def run(self, args: dict) -> dict:
        action = args.get("action", "check")
        path = args.get("file", "")
        code = args.get("code", "")
        try:
            if code and not path:
                with tempfile.NamedTemporaryFile(mode="w", suffix=".py", delete=False) as f:
                    f.write(code); path = f.name
            if action == "check":
                r = subprocess.run([sys.executable, "-m", "ruff", "check", path, "--output-format", "json"], capture_output=True, text=True)
                issues = json.loads(r.stdout) if r.stdout.strip() else []
                return {"issues": issues, "count": len(issues)}
            elif action == "fix":
                r = subprocess.run([sys.executable, "-m", "ruff", "check", path, "--fix"], capture_output=True, text=True)
                return {"output": r.stdout, "fixed": r.returncode == 0}
            elif action == "format":
                r = subprocess.run([sys.executable, "-m", "ruff", "format", path], capture_output=True, text=True)
                return {"output": r.stdout, "formatted": r.returncode == 0}
            return {"error": f"unknown: {action}"}
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),

    ("vector_store", "Vector Store", "Store and search by meaning. Free, local.", "chromadb",
'''#!/usr/bin/env python3
"""Vector Store - ChromaDB (Apache 2.0) - 100% Free"""
import json, sys, subprocess, os


class Skill:
    def __init__(self, config: dict = None):
        self.config = config or {}
        self.pdir = self.config.get("persist_dir", os.path.expanduser("~/.evolvix/chromadb"))

    def run(self, args: dict) -> dict:
        action = args.get("action", "search")
        col = args.get("collection", "evolvix")
        try:
            import chromadb
            client = chromadb.PersistentClient(path=args.get("persist_dir", self.pdir))
            collection = client.get_or_create_collection(col)
            if action == "add":
                docs = args.get("documents", [])
                ids = args.get("ids", [f"doc_{i}" for i in range(len(docs))])
                collection.add(documents=docs, ids=ids)
                return {"added": len(docs), "total": collection.count()}
            elif action == "search":
                results = collection.query(query_texts=[args.get("query","")], n_results=args.get("n_results",5))
                return {"results": results}
            elif action == "count":
                return {"count": collection.count()}
            elif action == "clear":
                client.delete_collection(col)
                return {"deleted": col}
            return {"error": f"unknown: {action}"}
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "-q", "chromadb"], capture_output=True)
            return self.run(args)
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),

    ("nlp_processor", "NLP Processor", "Extract entities, analyze text. Free, local.", "spacy",
'''#!/usr/bin/env python3
"""NLP Processor - spaCy (MIT) - 100% Free"""
import json, sys, subprocess


class Skill:
    def __init__(self, config: dict = None):
        self.config = config or {}
        self._nlp = None

    def run(self, args: dict) -> dict:
        text = args.get("text", "")
        if not text:
            return {"error": "text required"}
        try:
            if self._nlp is None:
                import spacy
                try:
                    self._nlp = spacy.load("en_core_web_sm")
                except:
                    subprocess.run([sys.executable, "-m", "spacy", "download", "en_core_web_sm"], capture_output=True)
                    self._nlp = spacy.load("en_core_web_sm")
            doc = self._nlp(text)
            return {
                "entities": [{"text": e.text, "label": e.label_} for e in doc.ents],
                "sentences": len(list(doc.sents)),
                "keywords": [t.lemma_ for t in doc if not t.is_stop and not t.is_punct][:20],
            }
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "-q", "spacy"], capture_output=True)
            return self.run(args)
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),

    ("math_solver", "Math Solver", "Solve equations, calculus, algebra. Free.", "sympy",
'''#!/usr/bin/env python3
"""Math Solver - SymPy (BSD) - 100% Free"""
import json, sys, subprocess


class Skill:
    def __init__(self, config: dict = None):
        self.config = config or {}

    def run(self, args: dict) -> dict:
        action = args.get("action", "solve")
        try:
            from sympy import symbols, solve, simplify, diff, integrate, sympify
            x = symbols("x")
            if action == "solve":
                eq = sympify(args["equation"])
                return {"solutions": [str(s) for s in solve(eq, x)]}
            elif action == "derive":
                expr = sympify(args["expression"])
                return {"derivative": str(simplify(diff(expr, x)))}
            elif action == "integrate":
                expr = sympify(args["expression"])
                return {"integral": str(integrate(expr, x))}
            elif action == "simplify":
                return {"simplified": str(simplify(sympify(args["expression"])))}
            elif action == "evaluate":
                return {"result": str(sympify(args["expression"]).evalf())}
            return {"error": f"unknown: {action}"}
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "-q", "sympy"], capture_output=True)
            return self.run(args)
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),

    ("ml_toolkit", "ML Toolkit", "Classification, clustering, PCA. Free, local.", "scikit-learn",
'''#!/usr/bin/env python3
"""ML Toolkit - scikit-learn (BSD) - 100% Free"""
import json, sys, subprocess


class Skill:
    def __init__(self, config: dict = None):
        self.config = config or {}

    def run(self, args: dict) -> dict:
        action = args.get("action", "classify")
        try:
            from sklearn import datasets, svm, cluster, decomposition
            from sklearn.model_selection import train_test_split
            from sklearn.metrics import accuracy_score
            import numpy as np
            if action == "classify":
                X = np.array(args.get("X", [])) if args.get("X") else datasets.load_iris().data
                y = np.array(args.get("y", [])) if args.get("y") else datasets.load_iris().target
                Xtr, Xte, ytr, yte = train_test_split(X, y, test_size=0.3)
                clf = svm.SVC().fit(Xtr, ytr)
                return {"accuracy": float(accuracy_score(yte, clf.predict(Xte)))}
            elif action == "cluster":
                X = np.array(args.get("X", [])) if args.get("X") else datasets.load_iris().data
                km = cluster.KMeans(n_clusters=args.get("n_clusters",3), n_init=10).fit(X)
                return {"labels": km.labels_.tolist(), "centers": km.cluster_centers_.tolist()}
            elif action == "pca":
                X = np.array(args.get("X", [])) if args.get("X") else datasets.load_iris().data
                pca = decomposition.PCA(n_components=args.get("n_components",2)).fit(X)
                return {"reduced": pca.transform(X).tolist(), "variance": pca.explained_variance_ratio_.tolist()}
            return {"error": f"unknown: {action}"}
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "-q", "scikit-learn"], capture_output=True)
            return self.run(args)
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),

    ("security_scanner_pro", "Security Scanner Pro", "Scan code for vulnerabilities. Free.", "semgrep bandit",
'''#!/usr/bin/env python3
"""Security Scanner Pro - Semgrep + Bandit (LGPL/Apache 2.0) - 100% Free"""
import json, sys, subprocess, os, tempfile


class Skill:
    def __init__(self, config: dict = None):
        self.config = config or {}

    def run(self, args: dict) -> dict:
        path = args.get("path", "")
        code = args.get("code", "")
        scanner = args.get("scanner", "bandit")
        try:
            if code and not path:
                with tempfile.NamedTemporaryFile(mode="w", suffix=".py", delete=False) as f:
                    f.write(code); path = f.name
            if scanner == "bandit":
                r = subprocess.run([sys.executable, "-m", "bandit", "-r", path, "-f", "json"], capture_output=True, text=True)
                data = json.loads(r.stdout) if r.stdout.strip() else {}
                return {"issues": data.get("results", []), "count": len(data.get("results", []))}
            elif scanner == "semgrep":
                r = subprocess.run(["semgrep", "--json", path], capture_output=True, text=True)
                data = json.loads(r.stdout) if r.stdout.strip() else {}
                return {"issues": data.get("results", []), "count": len(data.get("results", []))}
            return {"error": f"unknown scanner: {scanner}"}
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),

    ("scheduler_pro", "Scheduler Pro", "Schedule recurring tasks. Free, local.", "APScheduler",
'''#!/usr/bin/env python3
"""Scheduler Pro - APScheduler (MIT) - 100% Free"""
import json, sys, subprocess


class Skill:
    _scheduler = None

    def __init__(self, config: dict = None):
        self.config = config or {}

    def _get_sched(self):
        if Skill._scheduler is None:
            from apscheduler.schedulers.background import BackgroundScheduler
            Skill._scheduler = BackgroundScheduler()
            Skill._scheduler.start()
        return Skill._scheduler

    def run(self, args: dict) -> dict:
        action = args.get("action", "add")
        try:
            sched = self._get_sched()
            if action == "add":
                jtype = args.get("type", "interval")
                jid = args.get("id", f"job_{len(sched.get_jobs())}")
                if jtype == "interval":
                    sched.add_job(func=lambda: None, trigger="interval", seconds=args.get("seconds",60), id=jid)
                elif jtype == "cron":
                    sched.add_job(func=lambda: None, trigger="cron", hour=args.get("hour",0), minute=args.get("minute",0), id=jid)
                return {"status": "scheduled", "id": jid}
            elif action == "list":
                return {"jobs": [{"id": j.id, "next_run": str(j.next_run_time)} for j in sched.get_jobs()]}
            elif action == "remove":
                sched.remove_job(args["id"])
                return {"removed": args["id"]}
            return {"error": f"unknown: {action}"}
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "-q", "APScheduler"], capture_output=True)
            return self.run(args)
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),

    ("embeddings_engine", "Embeddings Engine", "Generate text embeddings. Free, local.", "sentence-transformers",
'''#!/usr/bin/env python3
"""Embeddings Engine - sentence-transformers (Apache 2.0) - 100% Free"""
import json, sys, subprocess


class Skill:
    _model = None

    def __init__(self, config: dict = None):
        self.config = config or {}

    def _get_model(self):
        if Skill._model is None:
            from sentence_transformers import SentenceTransformer
            Skill._model = SentenceTransformer("all-MiniLM-L6-v2")
        return Skill._model

    def run(self, args: dict) -> dict:
        action = args.get("action", "embed")
        try:
            model = self._get_model()
            if action == "embed":
                texts = args.get("texts", [])
                if isinstance(texts, str): texts = [texts]
                embs = model.encode(texts)
                return {"embeddings": embs.tolist(), "dim": embs.shape[1]}
            elif action == "similarity":
                e = model.encode([args["text1"], args["text2"]])
                score = float(e[0] @ e[1] / (e[0].norm() * e[1].norm()))
                return {"similarity": score}
            elif action == "search":
                docs = args.get("documents", [])
                qe = model.encode([args["query"]])
                de = model.encode(docs)
                scores = (de @ qe.T).flatten().tolist()
                ranked = sorted(zip(docs, scores), key=lambda x: x[1], reverse=True)
                return {"results": [{"text": t, "score": float(s)} for t, s in ranked[:args.get("top_k",5)]]}
            return {"error": f"unknown: {action}"}
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "-q", "sentence-transformers"], capture_output=True)
            return self.run(args)
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),

    ("dashboard_builder", "Dashboard Builder", "Build interactive dashboards. Free, local.", "streamlit",
'''#!/usr/bin/env python3
"""Dashboard Builder - Streamlit (Apache 2.0) - 100% Free"""
import json, sys, subprocess


class Skill:
    def __init__(self, config: dict = None):
        self.config = config or {}

    def run(self, args: dict) -> dict:
        action = args.get("action", "generate")
        try:
            if action == "generate":
                title = args.get("title", "EvolvixOS Dashboard")
                code_lines = [
                    "import streamlit as st",
                    "import pandas as pd",
                    'import plotly.express as px',
                    f'st.title("{title}")',
                    'file = st.file_uploader("Upload CSV", type=["csv"])',
                    "if file:",
                    "    df = pd.read_csv(file)",
                    "    st.dataframe(df)",
                    '    cols = df.select_dtypes(include="number").columns.tolist()',
                    "    if cols:",
                    '        x = st.selectbox("X", df.columns)',
                    '        y = st.selectbox("Y", cols)',
                    "        st.plotly_chart(px.bar(df, x=x, y=y))",
                ]
                out = args.get("output", "evolvix_dashboard.py")
                with open(out, "w") as f:
                    f.write("\\n".join(code_lines))
                return {"file": out, "run": f"streamlit run {out}"}
            elif action == "run":
                port = args.get("port", 8501)
                path = args.get("file", "evolvix_dashboard.py")
                subprocess.Popen([sys.executable, "-m", "streamlit", "run", path, "--server.port", str(port)])
                return {"status": "running", "port": port, "url": f"http://localhost:{port}"}
            return {"error": f"unknown: {action}"}
        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    args = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    print(json.dumps(Skill().run(args)))
'''),
]

for sid, name, desc, pip, code in SKILLS:
    d = os.path.join("skills", sid)
    os.makedirs(d, exist_ok=True)
    with open(os.path.join(d, "skill.py"), "w") as f:
        f.write(code)
    meta = {"name": name, "description": desc, "version": "1.0.0", "free": True, "local": True, "cloud": False, "cost": "$0.00", "pip_dependencies": pip}
    with open(os.path.join(d, "skill.json"), "w") as f:
        json.dump(meta, f, indent=2)
    with open(os.path.join(d, "__init__.py"), "w") as f:
        f.write(f"# {name}\n")
    print(f"✅ {sid}")

print(f"\n=== {len(SKILLS)+1} skills (incl web_crawler) ===")
